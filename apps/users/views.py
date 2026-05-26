"""NBES RBAC admin endpoints.
All endpoints require an authenticated user with the ``rbac:manage``
permission (held by ``system_administrator`` per the seed). Every mutation
emits an AuditEvent and invalidates the in-process role cache so the
change takes effect within 60 s for every NBES node.
"""

import logging

from django.db import transaction
from drf_spectacular.utils import (
    OpenApiExample,
    extend_schema,
    inline_serializer,
)
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import serializers
from rest_framework.views import APIView
from apps.audit.models import AuditEvent
from shared import rbac
from shared.permissions import has_permission, has_permission_with_step_up

from .models import (
    HIGH_PRIVILEGE_ROLES,
    Permission,
    Role,
    RoleAssignmentApproval,
    RoleChangeEvent,
    RoleMutualExclusion,
    RolePermission,
    UserRole,
)
from .serializers import (
    BulkRoleAssignSerializer,
    CreateRoleSerializer,
    PermissionSerializer,
    RoleApprovalActionSerializer,
    RoleAssignmentApprovalSerializer,
    RoleMutualExclusionCreateSerializer,
    RoleMutualExclusionSerializer,
    RoleSerializer,
    UpdateRolePermissionsSerializer,
    UserProfileCreateSerializer,
    UserProfileSerializer,
    UserProfileUpdateSerializer,
    UserRoleAssignSerializer,
    UserRoleSerializer,
)

logger = logging.getLogger(__name__)

# ── helpers ──────────────────────────────────────────────────────────────────


def _envelope(data, status_code=status.HTTP_200_OK, request_id=""):
    """NBES success envelope. Errors go through shared.exceptions handler."""
    return Response(
        {"success": True, "data": data, "meta": {"request_id": str(request_id)}},
        status=status_code,
    )


def _actor_id(request):
    return (request.auth or {}).get("sub") or None


def _audit(request, action, entity_id=None, new_state=None, old_state=None):
    AuditEvent.record(
        actor_id=_actor_id(request),
        action=action,
        entity_type="rbac",
        entity_id=entity_id,
        old_state=old_state,
        new_state=new_state,
        request_id=getattr(request, "request_id", None),
        ip_address=getattr(request, "ip_address", None),
    )


def _success_envelope(name, data_fields):
    return inline_serializer(
        name=name,
        fields={
            "success": serializers.BooleanField(default=True),
            "data": inline_serializer(name=f"{name}Data", fields=data_fields),
            "meta": inline_serializer(
                name=f"{name}Meta",
                fields={"request_id": serializers.CharField()},
            ),
        },
    )


def _success_envelope_with_serializer(name, data_serializer):
    return inline_serializer(
        name=name,
        fields={
            "success": serializers.BooleanField(default=True),
            "data": data_serializer,
            "meta": inline_serializer(
                name=f"{name}Meta",
                fields={"request_id": serializers.CharField()},
            ),
        },
    )


def _error_envelope(name):
    return inline_serializer(
        name=name,
        fields={
            "success": serializers.BooleanField(default=False),
            "error": inline_serializer(
                name=f"{name}Error",
                fields={
                    "code": serializers.CharField(),
                    "message": serializers.CharField(),
                },
            ),
            "meta": inline_serializer(
                name=f"{name}Meta",
                fields={"request_id": serializers.CharField()},
            ),
        },
    )


def _error_response(code, message, status_code, request_id=""):
    from shared.exceptions import format_rfc7807_error
    data = format_rfc7807_error(
        status_code=status_code,
        error_code=code,
        message=message,
        request_id=str(request_id),
    )
    return Response(data, status=status_code, content_type="application/problem+json")


# ── permissions catalog (read-only) ──────────────────────────────────────────


class PermissionListView(APIView):
    """``GET /api/v1/admin/rbac/permissions`` — list seeded codenames.

    Read-only: codenames are declared in code, never invented at runtime.
    """

    authentication_classes_setting = None  # use DRF default
    permission_classes = [IsAuthenticated, has_permission("rbac:manage")]

    @extend_schema(
        tags=["RBAC Admin"],
        summary="List NBES permission codenames",
        operation_id="rbac_permission_list",
        description="Returns the seeded permission catalog. Codenames are declared in code.",
        responses={
            200: _success_envelope(
                "PermissionListResponse",
                {
                    "count": serializers.IntegerField(),
                    "permissions": PermissionSerializer(many=True),
                },
            ),
            401: _error_envelope("UnauthorizedError"),
            403: _error_envelope("ForbiddenError"),
        },
    )
    def get(self, request):
        data = PermissionSerializer(
            Permission.objects.all().order_by("codename"), many=True
        ).data
        return _envelope(
            {"count": len(data), "permissions": data},
            request_id=getattr(request, "request_id", ""),
        )


# ── role registry (mirror of IAM roles NBES recognises) ─────────────────────


class RoleListCreateView(APIView):
    """``GET /api/v1/admin/rbac/roles`` — list registered roles.

    ``POST /api/v1/admin/rbac/roles`` — register an IAM role name so NBES
    starts recognising it. Does NOT create the role in Keycloak; that
    happens in IAM. NBES will only resolve permissions for roles whose
    names are in this table.
    """

    permission_classes = [IsAuthenticated, has_permission("rbac:manage")]

    @extend_schema(
        tags=["RBAC Admin"],
        summary="List NBES-recognised roles",
        operation_id="rbac_role_list",
        responses={
            200: _success_envelope(
                "RoleListResponse",
                {
                    "count": serializers.IntegerField(),
                    "roles": RoleSerializer(many=True),
                },
            ),
            401: _error_envelope("RoleListUnauthorizedError"),
            403: _error_envelope("RoleListForbiddenError"),
        },
    )
    def get(self, request):
        data = RoleSerializer(Role.objects.all().order_by("name"), many=True).data
        return _envelope(
            {"count": len(data), "roles": data},
            request_id=getattr(request, "request_id", ""),
        )

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Register an IAM role in NBES",
        operation_id="rbac_role_create",
        description=(
            "Creates or updates the local NBES role registry entry. "
            "It does not create the role in Keycloak."
        ),
        request=CreateRoleSerializer,
        responses={
            200: _success_envelope_with_serializer(
                "RoleUpdatedResponse", RoleSerializer()
            ),
            201: _success_envelope_with_serializer(
                "RoleCreatedResponse", RoleSerializer()
            ),
            400: _error_envelope("CreateRoleValidationError"),
            401: _error_envelope("CreateRoleUnauthorizedError"),
            403: _error_envelope("CreateRoleForbiddenError"),
        },
    )
    def post(self, request):
        serializer = CreateRoleSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data

        role, created = Role.objects.get_or_create(
            name=d["name"],
            defaults={"description": d.get("description", ""), "is_custom": True},
        )
        if not created and not role.is_custom:
            return _error_response(
                "ROLE_LOCKED",
                "Seeded NBES roles cannot be updated or deleted.",
                status.HTTP_403_FORBIDDEN,
                getattr(request, "request_id", ""),
            )

        old_description = role.description
        if not created and d.get("description") and not role.description:
            role.description = d["description"]
            role.save(update_fields=["description", "updated_at"])

        _audit(
            request,
            action="RBAC_ROLE_REGISTERED" if created else "RBAC_ROLE_UPDATED",
            entity_id=role.id,
            old_state=None
            if created
            else {"name": role.name, "description": old_description},
            new_state={"name": role.name, "description": role.description},
        )
        return _envelope(
            RoleSerializer(role).data,
            status_code=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
            request_id=getattr(request, "request_id", ""),
        )


class RoleDetailView(APIView):
    """``GET/PATCH/DELETE /api/v1/admin/rbac/roles/{id}``.

    DELETE only deactivates (``is_active=False``) so historical audit rows
    keep referring to a real role. The local cache is invalidated so
    revocations take effect within 60 s.
    """

    permission_classes = [IsAuthenticated, has_permission("rbac:manage")]

    def _get(self, pk):
        try:
            return Role.objects.get(id=pk), None
        except Role.DoesNotExist:
            return None, Response(
                {
                    "success": False,
                    "error": {"code": "NOT_FOUND", "message": "Role not found."},
                    "meta": {},
                },
                status=status.HTTP_404_NOT_FOUND,
            )

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Get a role",
        operation_id="rbac_role_retrieve",
        responses={
            200: _success_envelope_with_serializer(
                "RoleDetailResponse", RoleSerializer()
            ),
            404: _error_envelope("RoleNotFoundError"),
        },
    )
    def get(self, request, pk):
        role, err = self._get(pk)
        if err:
            return err
        return _envelope(
            RoleSerializer(role).data, request_id=getattr(request, "request_id", "")
        )

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Update a role",
        operation_id="rbac_role_update",
        request=inline_serializer(
            name="RolePatchRequest",
            fields={
                "description": serializers.CharField(required=False, allow_blank=True),
                "is_active": serializers.BooleanField(required=False),
            },
        ),
        responses={
            200: _success_envelope_with_serializer(
                "RolePatchResponse", RoleSerializer()
            ),
            404: _error_envelope("RolePatchNotFoundError"),
        },
    )
    def patch(self, request, pk):
        role, err = self._get(pk)
        if err:
            return err
        if not role.is_custom:
            return _error_response(
                "ROLE_LOCKED",
                "Seeded NBES roles cannot be updated or deleted.",
                status.HTTP_403_FORBIDDEN,
                getattr(request, "request_id", ""),
            )

        before = {"description": role.description, "is_active": role.is_active}
        if "description" in request.data:
            role.description = request.data["description"]
        if "is_active" in request.data:
            role.is_active = bool(request.data["is_active"])
        role.save(update_fields=["description", "is_active", "updated_at"])
        rbac.invalidate_role(role.name)

        _audit(
            request,
            action="RBAC_ROLE_UPDATED",
            entity_id=role.id,
            old_state=before,
            new_state={"description": role.description, "is_active": role.is_active},
        )
        return _envelope(
            RoleSerializer(role).data, request_id=getattr(request, "request_id", "")
        )

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Deactivate a role",
        operation_id="rbac_role_deactivate",
        responses={
            200: _success_envelope_with_serializer(
                "RoleDeactivateResponse",
                RoleSerializer(),
            ),
            404: _error_envelope("RoleDeactivateNotFoundError"),
        },
    )
    def delete(self, request, pk):
        role, err = self._get(pk)
        if err:
            return err
        if not role.is_custom:
            return _error_response(
                "ROLE_LOCKED",
                "Seeded NBES roles cannot be updated or deleted.",
                status.HTTP_403_FORBIDDEN,
                getattr(request, "request_id", ""),
            )

        if not role.is_active:
            return _envelope(
                {"detail": "Role is already inactive."},
                request_id=getattr(request, "request_id", ""),
            )

        role.is_active = False
        role.save(update_fields=["is_active", "updated_at"])
        rbac.invalidate_role(role.name)

        _audit(
            request,
            action="RBAC_ROLE_DEACTIVATED",
            entity_id=role.id,
            old_state={"is_active": True},
            new_state={"is_active": False},
        )
        return _envelope(
            RoleSerializer(role).data, request_id=getattr(request, "request_id", "")
        )


# ── role ↔ permission matrix (the editable bit) ─────────────────────────────


class RolePermissionsView(APIView):
    """``GET /api/v1/admin/rbac/roles/{id}/permissions`` — current grants.

    ``PUT /api/v1/admin/rbac/roles/{id}/permissions`` — replace with the
    given set. Computes additions/removals, persists in one transaction,
    invalidates the role cache.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("rbac:manage")]

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Get role permission grants",
        operation_id="rbac_role_permissions_retrieve",
        responses={
            200: _success_envelope(
                "RolePermissionsResponse",
                {
                    "role_id": serializers.UUIDField(),
                    "role_name": serializers.CharField(),
                    "permissions": serializers.ListField(child=serializers.CharField()),
                },
            ),
            404: _error_envelope("RolePermissionsNotFoundError"),
        },
    )
    def get(self, request, pk):
        try:
            role = Role.objects.get(id=pk)
        except Role.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "error": {"code": "NOT_FOUND", "message": "Role not found."},
                    "meta": {},
                },
                status=status.HTTP_404_NOT_FOUND,
            )
        codenames = sorted(role.grants.values_list("permission__codename", flat=True))
        return _envelope(
            {"role_id": str(role.id), "role_name": role.name, "permissions": codenames},
            request_id=getattr(request, "request_id", ""),
        )

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Replace role permission grants",
        operation_id="rbac_role_permissions_update",
        request=UpdateRolePermissionsSerializer,
        responses={
            200: _success_envelope(
                "RolePermissionsUpdateResponse",
                {
                    "role_id": serializers.UUIDField(),
                    "role_name": serializers.CharField(),
                    "permissions": serializers.ListField(child=serializers.CharField()),
                    "added": serializers.ListField(child=serializers.CharField()),
                    "removed": serializers.ListField(child=serializers.CharField()),
                },
            ),
            400: _error_envelope("RolePermissionsValidationError"),
            404: _error_envelope("RolePermissionsUpdateNotFoundError"),
        },
    )
    def put(self, request, pk):
        try:
            role = Role.objects.get(id=pk)
        except Role.DoesNotExist:
            return Response(
                {
                    "success": False,
                    "error": {"code": "NOT_FOUND", "message": "Role not found."},
                    "meta": {},
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        serializer = UpdateRolePermissionsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        desired = set(serializer.validated_data["codenames"])

        with transaction.atomic():
            current = set(role.grants.values_list("permission__codename", flat=True))
            to_remove = current - desired
            to_add = desired - current

            if to_remove:
                role.grants.filter(permission__codename__in=to_remove).delete()
            if to_add:
                permissions = Permission.objects.filter(codename__in=to_add)
                RolePermission.objects.bulk_create(
                    [
                        RolePermission(
                            role=role, permission=p, granted_by=_actor_id(request)
                        )
                        for p in permissions
                    ]
                )

            # increment version whenever the matrix changes
            if to_add or to_remove:
                role.version = (role.version or 1) + 1
                role.save(update_fields=["version", "updated_at"])

        rbac.invalidate_role(role.name)
        _audit(
            request,
            action="RBAC_ROLE_PERMISSIONS_UPDATED",
            entity_id=role.id,
            old_state={"permissions": sorted(current)},
            new_state={"permissions": sorted(desired), "version": role.version},
        )

        return _envelope(
            {
                "role_id": str(role.id),
                "role_name": role.name,
                "permissions": sorted(desired),
                "added": sorted(to_add),
                "removed": sorted(to_remove),
                "version": role.version,
            },
            request_id=getattr(request, "request_id", ""),
        )


# ── current-user introspection ──────────────────────────────────────────────


class MyPermissionsView(APIView):
    """``GET /api/v1/me/permissions`` — what NBES thinks the current user can do.

    Useful for clients to hide buttons before the API rejects them. The
    list reflects the *resolved* permissions after intersecting JWT roles
    with NBES's role registry — i.e. the same set the gateway enforces.

    **Role resolution (post-migration architecture):**

    1. NBES reads ``resource_access[<NBES_CLIENT_ID>].roles`` from the JWT —
       the Keycloak *client roles* IAM assigns when the user is invited /
       activated into NBES. ``NBES_CLIENT_ID`` defaults to ``nbes-api``.
    2. If that claim is absent (legacy tokens issued before IAM cut over),
       NBES falls back to ``realm_access.roles`` and emits a structured
       warning ``rbac.legacy_realm_role_fallback`` so the fallback usage is
       observable. The fallback will be removed in IAM Phase 7.
    3. ``super_admin`` in ``realm_access.roles`` short-circuits to the
       wildcard sentinel ``*``. This realm role stays a realm role; every
       other system role becomes a client role.

    **Audience verification.** Tokens for NBES must list ``nbes-api`` in
    ``aud`` (the ``audience-resolve`` mapper on ``clet-iam-internal``
    populates this automatically when the user holds NBES client roles).
    Production NBES (``prod.py``) fails closed at boot if
    ``KEYCLOAK_VALID_AUDIENCES`` is empty or missing ``nbes-api``.

    The ``roles_in_jwt`` field in the response reflects whichever source
    the resolver actually used — so an empty ``resource_access`` block on a
    legacy token will surface here as the realm-role names.
    """

    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Current User"],
        summary="Get current user's resolved NBES permissions",
        operation_id="current_user_permissions",
        description=(
            "Returns the resolved NBES permission set for the bearer of "
            "the JWT.\n\n"
            "**Resolution order:**\n"
            "1. `resource_access[nbes-api].roles` — Keycloak client roles "
            "(target architecture).\n"
            "2. `realm_access.roles` — legacy fallback; logged for "
            "migration tracking.\n"
            "3. `super_admin` in `realm_access.roles` → wildcard `*`.\n\n"
            "**Audience:** the token's `aud` claim must include the value "
            "of `NBES_CLIENT_ID` (default `nbes-api`)."
        ),
        responses={
            200: _success_envelope(
                "MyPermissionsResponse",
                {
                    "sub": serializers.CharField(allow_blank=True, required=False),
                    "email": serializers.EmailField(allow_blank=True, required=False),
                    "roles_in_jwt": serializers.ListField(
                        child=serializers.CharField()
                    ),
                    "roles_recognised_by_nbes": serializers.ListField(
                        child=serializers.CharField()
                    ),
                    "permissions": serializers.ListField(child=serializers.CharField()),
                },
            ),
            401: _error_envelope("MyPermissionsUnauthorizedError"),
        },
        examples=[
            OpenApiExample(
                name="Examiner — client-role path (target)",
                description=(
                    "User has `nbes-api/examiner` as a Keycloak client role. "
                    "Token's `resource_access['nbes-api'].roles` includes "
                    "`examiner`. NBES resolves `examiner` against its local "
                    "RolePermission table."
                ),
                value={
                    "success": True,
                    "data": {
                        "sub": "11111111-1111-1111-1111-111111111111",
                        "email": "examiner@example.com",
                        "roles_in_jwt": ["examiner"],
                        "roles_recognised_by_nbes": ["examiner"],
                        "permissions": ["marking:moderate", "marking:score"],
                    },
                    "meta": {"request_id": "0e1e..."},
                },
            ),
            OpenApiExample(
                name="Super admin — realm-role wildcard",
                description=(
                    "`super_admin` in `realm_access.roles` short-circuits "
                    "to wildcard. Token does not need `resource_access[nbes-api]`."
                ),
                value={
                    "success": True,
                    "data": {
                        "sub": "22222222-2222-2222-2222-222222222222",
                        "email": "root@example.com",
                        "roles_in_jwt": [],
                        "roles_recognised_by_nbes": [],
                        "permissions": ["*"],
                    },
                    "meta": {"request_id": "..."},
                },
            ),
            OpenApiExample(
                name="Legacy fallback — pre-migration token",
                description=(
                    "Token carries `realm_access.roles=['examiner']` but no "
                    "`resource_access[nbes-api]`. NBES still resolves it via "
                    "the realm-role fallback path AND emits "
                    "`rbac.legacy_realm_role_fallback` to logs."
                ),
                value={
                    "success": True,
                    "data": {
                        "sub": "33333333-3333-3333-3333-333333333333",
                        "email": "legacy@example.com",
                        "roles_in_jwt": ["examiner"],
                        "roles_recognised_by_nbes": ["examiner"],
                        "permissions": ["marking:moderate", "marking:score"],
                    },
                    "meta": {"request_id": "..."},
                },
            ),
        ],
    )
    def get(self, request):
        payload = request.auth or {}
        jwt_roles = rbac.get_nbes_role_names(payload)
        known_roles = sorted(
            Role.objects.filter(name__in=jwt_roles, is_active=True).values_list(
                "name", flat=True
            )
        )
        permissions = sorted(rbac.permissions_for(payload))
        return _envelope(
            {
                "sub": payload.get("sub"),
                "email": payload.get("email"),
                "roles_in_jwt": jwt_roles,
                "roles_recognised_by_nbes": known_roles,
                "permissions": permissions,
            },
            request_id=getattr(request, "request_id", ""),
        )


# ── role dashboard skeletons ────────────────────────────────────────────────


class DashboardView(APIView):
    """``GET /api/v1/me/dashboard`` — role dashboard skeleton for the current user.

    Delegates to the ``apps.dashboards`` DB-driven panel registry .
    Supports multi-role users: returns the union of all active panels for every
    role the bearer holds, ordered by (role_codename, display_order).
    """

    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Current User"],
        summary="Get role dashboard skeleton",
        operation_id="current_user_dashboard",
        responses={
            200: _success_envelope(
                "DashboardResponse",
                {
                    "roles": serializers.ListField(
                        child=serializers.CharField()
                    ),
                    "panels": serializers.ListField(
                        child=inline_serializer(
                            name="MeDashboardPanelItem",
                            fields={
                                "panel_key": serializers.CharField(),
                                "panel_name": serializers.CharField(),
                                "role_codename": serializers.CharField(),
                                "display_order": serializers.IntegerField(),
                                "default_config": serializers.JSONField(),
                            },
                        )
                    ),
                },
            ),
            401: _error_envelope("DashboardUnauthorizedError"),
        },
    )
    def get(self, request):
        from apps.dashboards.models import DashboardPanel

        payload = request.auth or {}
        roles = rbac.get_nbes_role_names(payload) or []

        panels_qs = (
            DashboardPanel.objects
            .filter(role_codename__in=roles, is_active=True)
            .order_by("role_codename", "display_order", "panel_name")
            .values(
                "panel_key", "panel_name", "role_codename",
                "display_order", "default_config",
            )
        )

        return _envelope(
            {"roles": sorted(set(roles)), "panels": list(panels_qs)},
            request_id=getattr(request, "request_id", ""),
        )


# ── user profile administration CRUD ─────────────────────────────────────────


class AdminUserListCreateView(APIView):
    """``GET /api/v1/admin/users`` — list and filter local profiles.

    ``POST /api/v1/admin/users`` — create a profile, provision in Keycloak,
    and map roles.
    """

    permission_classes = [IsAuthenticated, has_permission("users:manage")]

    @extend_schema(
        tags=["User Administration"],
        summary="List and filter user profiles",
        operation_id="admin_users_list",
        responses={
            200: _success_envelope(
                "UserListResponse",
                {
                    "count": serializers.IntegerField(),
                    "users": UserProfileSerializer(many=True),
                },
            ),
            401: _error_envelope("UserListUnauthorizedError"),
            403: _error_envelope("UserListForbiddenError"),
        },
    )
    def get(self, request):
        from apps.users.models import UserProfile
        from django.db.models import Q
        from shared.pagination import StandardResultsPagination

        qs = UserProfile.objects.all().order_by("email")

        # Filters
        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)

        role_filter = request.query_params.get("role")
        if role_filter:
            qs = qs.filter(
                user_roles__role__name=role_filter, user_roles__revoked_at__isnull=True
            )

        search_filter = request.query_params.get("search")
        if search_filter:
            qs = qs.filter(
                Q(first_name__icontains=search_filter)
                | Q(last_name__icontains=search_filter)
                | Q(email__icontains=search_filter)
            ).distinct()

        paginator = StandardResultsPagination()
        page = paginator.paginate_queryset(qs, request)
        serializer = UserProfileSerializer(page, many=True)
        return paginator.get_paginated_response(serializer.data)

    @extend_schema(
        tags=["User Administration"],
        summary="Create user profile and provision in Keycloak",
        operation_id="admin_user_create",
        request=UserProfileCreateSerializer,
        responses={
            201: _success_envelope_with_serializer(
                "UserCreatedResponse", UserProfileSerializer()
            ),
            400: _error_envelope("UserCreateValidationError"),
            401: _error_envelope("UserCreateUnauthorizedError"),
            403: _error_envelope("UserCreateForbiddenError"),
        },
    )
    def post(self, request):
        from apps.users.services import create_user, ServiceError

        serializer = UserProfileCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            user = create_user(
                email=serializer.validated_data["email"],
                first_name=serializer.validated_data["first_name"],
                last_name=serializer.validated_data["last_name"],
                roles=serializer.validated_data["roles"],
                effective_date=serializer.validated_data.get("effective_date"),
                metadata=serializer.validated_data.get("metadata", {}),
                created_by=request.user,
                actor_id=_actor_id(request),
                request_id=getattr(request, "request_id", None),
                ip_address=getattr(request, "ip_address", None),
            )
        except ServiceError as exc:
            return _error_response(
                exc.code, str(exc),
                status.HTTP_400_BAD_REQUEST,
                getattr(request, "request_id", ""),
            )

        return _envelope(
            UserProfileSerializer(user).data,
            status_code=status.HTTP_201_CREATED,
            request_id=getattr(request, "request_id", ""),
        )


class AdminUserDetailView(APIView):
    """``GET /api/v1/admin/users/{id}`` — retrieve user profile.

    ``PATCH /api/v1/admin/users/{id}`` — update, deactivate or logically delete profile.
    """

    permission_classes = [IsAuthenticated, has_permission("users:manage")]

    @extend_schema(
        tags=["User Administration"],
        summary="Retrieve a user profile",
        operation_id="admin_user_retrieve",
        responses={
            200: _success_envelope_with_serializer(
                "UserDetailResponse", UserProfileSerializer()
            ),
            401: _error_envelope("UserRetrieveUnauthorizedError"),
            403: _error_envelope("UserRetrieveForbiddenError"),
            404: _error_envelope("UserRetrieveNotFoundError"),
        },
    )
    def get(self, request, pk):
        from apps.users.models import UserProfile

        try:
            user = UserProfile.objects.get(pk=pk)
        except UserProfile.DoesNotExist:
            return _error_response(
                "NOT_FOUND",
                "User profile not found.",
                status.HTTP_404_NOT_FOUND,
                getattr(request, "request_id", ""),
            )
        return _envelope(
            UserProfileSerializer(user).data,
            request_id=getattr(request, "request_id", ""),
        )

    @extend_schema(
        tags=["User Administration"],
        summary="Update, deactivate or delete a user profile",
        operation_id="admin_user_update",
        request=UserProfileUpdateSerializer,
        responses={
            200: _success_envelope_with_serializer(
                "UserUpdatedResponse", UserProfileSerializer()
            ),
            400: _error_envelope("UserUpdateValidationError"),
            401: _error_envelope("UserUpdateUnauthorizedError"),
            403: _error_envelope("UserUpdateForbiddenError"),
            404: _error_envelope("UserUpdateNotFoundError"),
        },
    )
    def patch(self, request, pk):
        from apps.users.models import UserProfile, UserRole
        from shared import keycloak_admin
        from django.utils import timezone
        from shared.events import publish

        serializer_data = request.data

        with transaction.atomic():
            try:
                user = UserProfile.objects.select_for_update().get(pk=pk)
            except UserProfile.DoesNotExist:
                return _error_response(
                    "NOT_FOUND",
                    "User profile not found.",
                    status.HTTP_404_NOT_FOUND,
                    getattr(request, "request_id", ""),
                )

            serializer = UserProfileUpdateSerializer(user, data=serializer_data, partial=True)
            serializer.is_valid(raise_exception=True)

            old_state = {
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "status": user.status,
                "metadata": user.metadata,
            }

            # Check deactivation/delete
            deactivate = False
            if (
                serializer.validated_data.get("status") == "inactive"
                and user.status != "inactive"
            ):
                deactivate = True
            if (
                serializer.validated_data.get("deleted") is True
                and user.status != "inactive"
            ):
                deactivate = True

            if deactivate:
                from apps.users.services import has_active_assignments
                blocking_reasons = has_active_assignments(user)
                if blocking_reasons:
                    reasons_str = " and ".join(blocking_reasons)
                    return _error_response(
                        "USER_HAS_ACTIVE_ASSIGNMENTS",
                        f"Cannot deactivate this profile: user has {reasons_str}. "
                        "Reassign or close them before deactivating.",
                        status.HTTP_400_BAD_REQUEST,
                        getattr(request, "request_id", ""),
                    )

                user.status = "inactive"
                user.deactivated_at = timezone.now()

                # Schedule IAM deactivation after DB commit so the external call
                # does not hold the DB lock open.
                if user.keycloak_sub:
                    _keycloak_sub = str(user.keycloak_sub)

                    def _deactivate_in_iam():
                        try:
                            keycloak_admin.deactivate_user(_keycloak_sub)
                        except keycloak_admin.IntegrationError as exc:
                            import logging as _log
                            _log.getLogger(__name__).warning(
                                "patch: IAM deactivation failed sub=%s: %s",
                                _keycloak_sub, exc,
                            )

                    transaction.on_commit(_deactivate_in_iam)

            # Update other fields
            for field in ["first_name", "last_name", "email", "metadata"]:
                if field in serializer.validated_data:
                    setattr(user, field, serializer.validated_data[field])

            user.save()

            new_state = {
                "first_name": user.first_name,
                "last_name": user.last_name,
                "email": user.email,
                "status": user.status,
                "metadata": user.metadata,
            }

            _audit(
                request,
                action="USER_UPDATED",
                entity_id=user.id,
                old_state=old_state,
                new_state=new_state,
            )

            publish(
                "UserUpdated",
                {
                    "user_id": str(user.id),
                    "keycloak_sub": str(user.keycloak_sub) if user.keycloak_sub else None,
                    "email": user.email,
                    "status": user.status,
                },
            )

        return _envelope(
            UserProfileSerializer(user).data,
            request_id=getattr(request, "request_id", ""),
        )


class MyProfileView(APIView):
    """``GET /api/v1/me`` — current user profile, active roles, and resolved permissions."""

    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=["Current User"],
        summary="Get current user's profile, active roles and resolved permissions",
        operation_id="current_user_profile",
        responses={
            200: _success_envelope(
                "MyProfileResponse",
                {
                    "id": serializers.UUIDField(),
                    "email": serializers.EmailField(),
                    "first_name": serializers.CharField(),
                    "last_name": serializers.CharField(),
                    "status": serializers.CharField(),
                    "metadata": serializers.JSONField(),
                    "roles": UserRoleSerializer(many=True),
                    "effective_permissions": serializers.ListField(
                        child=serializers.CharField()
                    ),
                },
            ),
            401: _error_envelope("MyProfileUnauthorizedError"),
        },
    )
    def get(self, request):
        from apps.users.models import UserProfile, UserRole, RolePermission

        payload = request.auth or {}
        sub = payload.get("sub")

        user = request.user
        if not isinstance(user, UserProfile):
            user = UserProfile.objects.filter(keycloak_sub=sub).first()
            if not user:
                return _error_response(
                    "NOT_FOUND",
                    "User profile not found.",
                    status.HTTP_404_NOT_FOUND,
                    getattr(request, "request_id", ""),
                )

        # Resolve active permissions canonically (unions JWT claims + DB assignments)
        from shared.rbac import permissions_for
        effective_permissions = sorted(list(permissions_for(payload)))

        data = UserProfileSerializer(user).data
        data["effective_permissions"] = effective_permissions

        return _envelope(
            data,
            request_id=getattr(request, "request_id", ""),
        )


# ── Role Assignment / Revocation (POST /admin/users/{id}/roles/) ─────────────
# Addresses (mutual exclusion check) and (two-admin approval)


class AdminUserRolesView(APIView):
    """``GET /api/v1/admin/users/{id}/roles/`` — list current role assignments.

    ``POST /api/v1/admin/users/{id}/roles/`` — assign or revoke a role.

    Assign flow:
      1. Validate the requested role exists and is active.
      2. Run mutual-exclusion check — 409 if conflict.
      3. If the role is in ``HIGH_PRIVILEGE_ROLES`` — create a pending
         ``RoleAssignmentApproval`` and return HTTP 202.
      4. Otherwise create the ``UserRole`` directly, emit ``RoleChangeEvent``,
         call IAM, and publish the outbox event.

    Revoke flow:
      Mark the active ``UserRole`` as revoked, emit event, call IAM.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("users:manage")]

    @extend_schema(
        tags=["User Administration"],
        summary="List role assignments for a user",
        operation_id="admin_user_roles_list",
        responses={
            200: _success_envelope(
                "UserRolesListResponse",
                {
                    "count": serializers.IntegerField(),
                    "roles": UserRoleSerializer(many=True),
                },
            ),
            404: _error_envelope("UserRolesListNotFoundError"),
        },
    )
    def get(self, request, pk):
        from apps.users.models import UserProfile
        try:
            user = UserProfile.objects.get(pk=pk)
        except UserProfile.DoesNotExist:
            return _error_response(
                "NOT_FOUND",
                "User profile not found.",
                status.HTTP_404_NOT_FOUND,
                getattr(request, "request_id", ""),
            )
        active_roles = user.user_roles.filter(revoked_at__isnull=True).select_related(
            "role"
        )
        data = UserRoleSerializer(active_roles, many=True).data
        return _envelope(
            {"count": len(data), "roles": data},
            request_id=getattr(request, "request_id", ""),
        )

    @extend_schema(
        tags=["User Administration"],
        summary="Assign or revoke a role on a user profile",
        operation_id="admin_user_roles_assign",
        description=(
            "Assign: validates mutual-exclusion rules (409 if conflict). "
            "High-privilege roles (director_general, system_administrator, nbec_member) "
            "enter a two-administrator approval workflow — returns 202 with a pending "
            "RoleAssignmentApproval ID.\n\n"
            "Revoke: immediately revokes the active role assignment."
        ),
        request=UserRoleAssignSerializer,
        responses={
            200: _success_envelope_with_serializer(
                "UserRoleAssignResponse", UserRoleSerializer()
            ),
            202: _success_envelope(
                "UserRoleApprovalPendingResponse",
                {
                    "approval_id": serializers.UUIDField(),
                    "message": serializers.CharField(),
                    "expires_at": serializers.DateTimeField(),
                },
            ),
            400: _error_envelope("UserRoleAssignValidationError"),
            404: _error_envelope("UserRoleAssignNotFoundError"),
            409: _error_envelope("UserRoleAssignConflictError"),
        },
    )
    def post(self, request, pk):
        from apps.users.models import UserProfile
        from apps.users.services import assign_role, revoke_role, ServiceError, RoleApprovalPending

        try:
            user = UserProfile.objects.get(pk=pk)
        except UserProfile.DoesNotExist:
            return _error_response(
                "NOT_FOUND", "User profile not found.",
                status.HTTP_404_NOT_FOUND, getattr(request, "request_id", ""),
            )

        serializer = UserRoleAssignSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        action = serializer.validated_data["action"]
        role_name = serializer.validated_data["role"]
        reason = serializer.validated_data.get("reason", "")
        effective_from = serializer.validated_data.get("effective_from")
        effective_to = serializer.validated_data.get("effective_to")

        common = dict(
            actor_id=_actor_id(request),
            request_id=getattr(request, "request_id", None),
            ip_address=getattr(request, "ip_address", None),
        )

        if action == UserRoleAssignSerializer.ACTION_ASSIGN:
            try:
                user_role = assign_role(
                    user=user,
                    role_name=role_name,
                    effective_from=effective_from,
                    effective_to=effective_to,
                    assigned_by=request.user,
                    reason=reason,
                    **common,
                )
            except RoleApprovalPending as exc:
                approval = exc.approval
                return Response(
                    {
                        "success": True,
                        "data": {
                            "approval_id": str(approval.id),
                            "message": (
                                f"Role '{role_name}' requires two-administrator "
                                "approval. A second administrator must approve "
                                "before the role is granted."
                            ),
                            "expires_at": approval.expires_at,
                        },
                        "meta": {"request_id": str(getattr(request, "request_id", ""))},
                    },
                    status=status.HTTP_202_ACCEPTED,
                )
            except ServiceError as exc:
                http_status = (
                    status.HTTP_409_CONFLICT
                    if exc.code in ("ROLE_CONFLICT", "ROLE_ALREADY_ASSIGNED")
                    else status.HTTP_404_NOT_FOUND
                    if exc.code == "ROLE_NOT_FOUND"
                    else status.HTTP_400_BAD_REQUEST
                )
                return _error_response(
                    exc.code, str(exc), http_status,
                    getattr(request, "request_id", ""),
                )
            return _envelope(
                UserRoleSerializer(user_role).data,
                request_id=getattr(request, "request_id", ""),
            )

        else:  # revoke
            try:
                revoke_role(
                    user=user,
                    role_name=role_name,
                    reason=reason,
                    actor=request.user,
                    **common,
                )
            except ServiceError as exc:
                return _error_response(
                    exc.code, str(exc),
                    status.HTTP_404_NOT_FOUND,
                    getattr(request, "request_id", ""),
                )
            return _envelope(
                {"detail": f"Role '{role_name}' revoked."},
                request_id=getattr(request, "request_id", ""),
            )


# ── Mutual-Exclusion Rule Administration ─────────────────────────────────────


class RoleMutualExclusionListCreateView(APIView):
    """``GET /api/v1/admin/rbac/exclusions/`` — list all configured exclusion pairs.

    ``POST /api/v1/admin/rbac/exclusions/`` — define a new exclusion pair.
    Roles are sorted server-side so (A, B) and (B, A) are idempotent.
    Requires ``rbac:manage``.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("rbac:manage")]

    @extend_schema(
        tags=["RBAC Admin"],
        summary="List mutual-exclusion rules",
        operation_id="rbac_exclusions_list",
        responses={
            200: _success_envelope(
                "ExclusionListResponse",
                {
                    "count": serializers.IntegerField(),
                    "exclusions": RoleMutualExclusionSerializer(many=True),
                },
            ),
        },
    )
    def get(self, request):
        qs = RoleMutualExclusion.objects.select_related("role_a", "role_b").all()
        data = RoleMutualExclusionSerializer(qs, many=True).data
        return _envelope(
            {"count": len(data), "exclusions": data},
            request_id=getattr(request, "request_id", ""),
        )

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Create a mutual-exclusion rule",
        operation_id="rbac_exclusions_create",
        request=RoleMutualExclusionCreateSerializer,
        responses={
            201: _success_envelope_with_serializer(
                "ExclusionCreatedResponse", RoleMutualExclusionSerializer()
            ),
            400: _error_envelope("ExclusionCreateValidationError"),
        },
    )
    def post(self, request):
        serializer = RoleMutualExclusionCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        d = serializer.validated_data
        actor_id = _actor_id(request)
        exclusion = RoleMutualExclusion.objects.create(
            role_a=d["role_a_obj"],
            role_b=d["role_b_obj"],
            reason=d.get("reason", ""),
            created_by=actor_id,
        )
        _audit(
            request,
            action="RBAC_EXCLUSION_CREATED",
            entity_id=exclusion.id,
            new_state={
                "role_a": d["role_a_obj"].name,
                "role_b": d["role_b_obj"].name,
                "reason": d.get("reason", ""),
            },
        )
        return _envelope(
            RoleMutualExclusionSerializer(exclusion).data,
            status_code=status.HTTP_201_CREATED,
            request_id=getattr(request, "request_id", ""),
        )


class RoleMutualExclusionDetailView(APIView):
    """``DELETE /api/v1/admin/rbac/exclusions/{id}/`` — remove an exclusion rule.

    Once deleted, the two roles can coexist again. Requires ``rbac:manage``.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("rbac:manage")]

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Delete a mutual-exclusion rule",
        operation_id="rbac_exclusions_delete",
        responses={
            200: _success_envelope(
                "ExclusionDeletedResponse", {"detail": serializers.CharField()}
            ),
            404: _error_envelope("ExclusionNotFoundError"),
        },
    )
    def delete(self, request, pk):
        try:
            exclusion = RoleMutualExclusion.objects.select_related(
                "role_a", "role_b"
            ).get(pk=pk)
        except RoleMutualExclusion.DoesNotExist:
            return _error_response(
                "NOT_FOUND",
                "Exclusion rule not found.",
                status.HTTP_404_NOT_FOUND,
                getattr(request, "request_id", ""),
            )
        names = (exclusion.role_a.name, exclusion.role_b.name)
        _audit(
            request,
            action="RBAC_EXCLUSION_DELETED",
            entity_id=exclusion.id,
            old_state={"role_a": names[0], "role_b": names[1]},
        )
        exclusion.delete()
        return _envelope(
            {
                "detail": f"Exclusion rule between '{names[0]}' and '{names[1]}' removed."
            },
            request_id=getattr(request, "request_id", ""),
        )


# ── Two-Administrator Approval Workflow ───────────────────────────────────────


class RoleAssignmentApprovalListView(APIView):
    """``GET /api/v1/admin/rbac/approvals/`` — list pending approvals.

    Supports ``?status=pending|approved|rejected|expired`` filter.
    Requires ``rbac:manage``.
    """

    permission_classes = [IsAuthenticated, has_permission("rbac:manage")]

    @extend_schema(
        tags=["RBAC Admin"],
        summary="List role-assignment approvals",
        operation_id="rbac_approvals_list",
        responses={
            200: _success_envelope(
                "ApprovalListResponse",
                {
                    "count": serializers.IntegerField(),
                    "approvals": RoleAssignmentApprovalSerializer(many=True),
                },
            ),
        },
    )
    def get(self, request):
        qs = RoleAssignmentApproval.objects.select_related(
            "target_user", "role", "requested_by", "reviewed_by"
        ).all()
        status_filter = request.query_params.get("status")
        if status_filter:
            qs = qs.filter(status=status_filter)
        data = RoleAssignmentApprovalSerializer(qs, many=True).data
        return _envelope(
            {"count": len(data), "approvals": data},
            request_id=getattr(request, "request_id", ""),
        )


class RoleAssignmentApprovalActionView(APIView):
    """``POST /api/v1/admin/rbac/approvals/{id}/approve/`` — approve a pending request.

    ``POST /api/v1/admin/rbac/approvals/{id}/reject/`` — reject a pending request.

    The reviewing administrator must differ from the requesting administrator.
    On approval, the ``UserRole`` is created atomically and Keycloak is updated.
    Requires ``rbac:manage``.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("rbac:manage")]

    def _get_approval(self, pk):
        try:
            return RoleAssignmentApproval.objects.select_related(
                "target_user", "role", "requested_by"
            ).get(pk=pk)
        except RoleAssignmentApproval.DoesNotExist:
            return None

    @extend_schema(
        tags=["RBAC Admin"],
        summary="Approve a pending role-assignment request",
        operation_id="rbac_approvals_approve",
        request=RoleApprovalActionSerializer,
        responses={
            200: _success_envelope_with_serializer(
                "ApprovalApprovedResponse", RoleAssignmentApprovalSerializer()
            ),
            400: _error_envelope("ApprovalActionError"),
            403: _error_envelope("ApprovalForbiddenError"),
            404: _error_envelope("ApprovalNotFoundError"),
        },
    )
    @transaction.atomic
    def post(self, request, pk, action):
        from shared import keycloak_admin
        from shared.events import publish

        approval = self._get_approval(pk)
        if not approval:
            return _error_response(
                "NOT_FOUND",
                "Approval request not found.",
                status.HTTP_404_NOT_FOUND,
                getattr(request, "request_id", ""),
            )

        serializer = RoleApprovalActionSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        note = serializer.validated_data.get("note", "")

        reviewer = request.user
        try:
            if action == "approve":
                user_role = approval.do_approve(reviewer, note)
                # Assign client role in Keycloak
                if approval.target_user.keycloak_sub:
                    try:
                        keycloak_admin.assign_client_role(
                            str(approval.target_user.keycloak_sub), approval.role.name
                        )
                    except keycloak_admin.IntegrationError:
                        logger.warning(
                            "approval: keycloak assign_client_role failed user=%s role=%s"
                            " — DB approved; IAM will sync on next login",
                            approval.target_user.id, approval.role.name,
                        )
                _audit(
                    request,
                    action="ROLE_APPROVAL_APPROVED",
                    entity_id=approval.target_user.id,
                    new_state={
                        "role": approval.role.name,
                        "approval_id": str(approval.id),
                        "reviewer": reviewer.email,
                    },
                )
                publish(
                    "UserRoleChanged",
                    {
                        "user_id": str(approval.target_user.id),
                        "change_type": "assign",
                        "role": approval.role.name,
                        "via_approval": str(approval.id),
                    },
                )
            else:  # reject
                approval.do_reject(reviewer, note)
                _audit(
                    request,
                    action="ROLE_APPROVAL_REJECTED",
                    entity_id=approval.target_user.id,
                    old_state={
                        "role": approval.role.name,
                        "approval_id": str(approval.id),
                    },
                    new_state={"reviewer": reviewer.email, "note": note},
                )
        except ValueError as exc:
            return _error_response(
                "APPROVAL_ACTION_INVALID",
                str(exc),
                status.HTTP_400_BAD_REQUEST
                if "status" in str(exc)
                else status.HTTP_403_FORBIDDEN,
                getattr(request, "request_id", ""),
            )

        return _envelope(
            RoleAssignmentApprovalSerializer(approval).data,
            request_id=getattr(request, "request_id", ""),
        )


# ── Bulk User Import  ─────────────────────────────────────────────


class BulkUserImportView(APIView):
    """``POST /api/v1/admin/users/import`` — import users from CSV or XLSX.

    Blueprint §1.2.4:
    * Valid rows are provisioned in IAM and stored as local profiles.
    * Invalid rows are reported with row-level errors (partial-success).
    * The uploaded file's SHA-256 hash is recorded; the file is retained
      for 7 years per the compliance requirement.
    * Supported content types: ``text/csv``, ``application/vnd.openxmlformats-
      officedocument.spreadsheetml.sheet`` (.xlsx).

    Required CSV/XLSX columns (case-insensitive):
      ``first_name``, ``last_name``, ``email``, ``roles``
    Optional columns:
      ``national_id``, ``department``
    The ``roles`` column accepts a comma-separated list of role names.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("users:import")]

    from rest_framework.parsers import MultiPartParser, FormParser
    parser_classes = [MultiPartParser, FormParser]

    @extend_schema(
        tags=["User Administration"],
        summary="Bulk import users from CSV or XLSX",
        operation_id="admin_users_bulk_import",
        request={
            "multipart/form-data": inline_serializer(
                name="BulkImportRequest",
                fields={"file": serializers.FileField()},
            )
        },
        responses={
            200: _success_envelope_with_serializer(
                "BulkImportResponse",
                inline_serializer(
                    name="BulkImportResponseData",
                    fields={
                        "import_id": serializers.UUIDField(),
                        "total_rows": serializers.IntegerField(),
                        "success_count": serializers.IntegerField(),
                        "failure_count": serializers.IntegerField(),
                        "row_errors": serializers.ListField(
                            child=serializers.JSONField()
                        ),
                    },
                ),
            ),
            400: _error_envelope("BulkImportValidationError"),
            401: _error_envelope("BulkImportUnauthorizedError"),
            403: _error_envelope("BulkImportForbiddenError"),
        },
    )
    def post(self, request):
        import csv
        import hashlib
        import io
        from django.utils import timezone as tz

        from apps.users.models import (
            BulkImportRecord,
            Role,
            RoleChangeEvent,
            UserProfile,
            UserRole,
        )
        from shared import keycloak_admin
        from shared.events import publish

        uploaded = request.FILES.get("file")
        if not uploaded:
            return _error_response(
                "MISSING_FILE",
                "No file uploaded.",
                status.HTTP_400_BAD_REQUEST,
                getattr(request, "request_id", ""),
            )

        filename = uploaded.name.lower()
        raw = uploaded.read()
        file_hash = hashlib.sha256(raw).hexdigest()

        # ── Parse rows ────────────────────────────────────────────────────────
        rows = []
        parse_error = None
        if filename.endswith(".csv"):
            try:
                text = raw.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
            except Exception as exc:
                parse_error = str(exc)
        elif filename.endswith(".xlsx"):
            try:
                import openpyxl

                wb = openpyxl.load_workbook(
                    io.BytesIO(raw), read_only=True, data_only=True
                )
                ws = wb.active
                headers = [
                    str(c.value).strip().lower() if c.value else ""
                    for c in next(ws.iter_rows())
                ]
                for excel_row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(
                        {headers[i]: (v or "") for i, v in enumerate(excel_row)}
                    )
            except Exception as exc:
                parse_error = str(exc)
        else:
            return _error_response(
                "UNSUPPORTED_FILE_TYPE",
                "Only .csv and .xlsx files are supported.",
                status.HTTP_400_BAD_REQUEST,
                getattr(request, "request_id", ""),
            )

        if parse_error:
            return _error_response(
                "FILE_PARSE_ERROR",
                f"Could not parse file: {parse_error}",
                status.HTTP_400_BAD_REQUEST,
                getattr(request, "request_id", ""),
            )

        REQUIRED_COLS = {"first_name", "last_name", "email", "roles"}
        if rows:
            col_keys = {k.strip().lower() for k in rows[0].keys()}
            missing = REQUIRED_COLS - col_keys
            if missing:
                return _error_response(
                    "MISSING_COLUMNS",
                    f"File is missing required columns: {sorted(missing)}",
                    status.HTTP_400_BAD_REQUEST,
                    getattr(request, "request_id", ""),
                )

        today = tz.now().date()
        record = BulkImportRecord.objects.create(
            uploaded_by=request.user,
            original_filename=uploaded.name,
            file_hash=file_hash,
            status=BulkImportRecord.STATUS_PROCESSING,
            total_rows=len(rows),
        )

        # ── Process each row (partial-success) ────────────────────────────────
        successes = 0
        row_errors = []

        for i, raw_row in enumerate(rows, start=2):  # row 2 = first data row
            row = {k.strip().lower(): ("" if v is None else str(v)).strip() for k, v in raw_row.items()}
            email = row.get("email", "")
            errs = []

            # Validate
            if not email:
                errs.append("email is required")
            if not row.get("first_name"):
                errs.append("first_name is required")
            if not row.get("last_name"):
                errs.append("last_name is required")
            role_names = [
                r.strip() for r in row.get("roles", "").split(",") if r.strip()
            ]
            if not role_names:
                errs.append("at least one role is required")

            if not errs and UserProfile.objects.filter(email__iexact=email).exists():
                errs.append(f"email '{email}' already exists")

            if not errs:
                valid_roles = list(
                    Role.objects.filter(
                        name__in=role_names, is_active=True
                    ).values_list("name", flat=True)
                )
                unknown_roles = set(role_names) - set(valid_roles)
                if unknown_roles:
                    errs.append(f"unknown roles: {sorted(unknown_roles)}")

            if errs:
                row_errors.append({"row": i, "email": email, "errors": errs})
                continue

            # Provision in IAM
            try:
                iam_sub = keycloak_admin.create_user(
                    email=email,
                    first_name=row["first_name"],
                    last_name=row["last_name"],
                    roles=valid_roles,
                    send_invite=True,
                )
            except Exception as exc:
                row_errors.append(
                    {"row": i, "email": email, "errors": [f"IAM error: {exc}"]}
                )
                continue

            keycloak_sub_val = None
            try:
                import uuid as _uuid

                keycloak_sub_val = _uuid.UUID(iam_sub)
            except (ValueError, TypeError):
                pass

            metadata = {}
            if row.get("national_id"):
                metadata["national_id"] = row["national_id"]
            if row.get("department"):
                metadata["department"] = row["department"]

            try:
                with transaction.atomic():
                    user = UserProfile.objects.create(
                        keycloak_sub=keycloak_sub_val,
                        email=email,
                        first_name=row["first_name"],
                        last_name=row["last_name"],
                        status="pending_invite",
                        metadata=metadata,
                        created_by=request.user,
                    )
                    for role_name in valid_roles:
                        role_obj = Role.objects.get(name=role_name)
                        UserRole.objects.create(
                            user=user,
                            role=role_obj,
                            effective_from=today,
                            assigned_by=request.user,
                        )
                        RoleChangeEvent.objects.create(
                            user=user,
                            role=role_obj,
                            change_type="assign",
                            actor=request.user,
                            reason="Bulk import",
                        )
                    publish(
                        "UserCreated",
                        {
                            "user_id": str(user.id),
                            "keycloak_sub": str(keycloak_sub_val) if keycloak_sub_val else None,
                            "email": email,
                            "roles": valid_roles,
                            "source": "bulk_import",
                        },
                    )
            except Exception as exc:
                # IAM user was already provisioned; roll it back so we don't
                # leave an orphaned Keycloak account with no local profile.
                if iam_sub:
                    try:
                        keycloak_admin.deactivate_user(iam_sub)
                    except Exception as iam_exc:
                        logger.warning(
                            "bulk_import: IAM rollback failed for sub=%s: %s",
                            iam_sub, iam_exc,
                        )
                row_errors.append(
                    {"row": i, "email": email, "errors": [f"DB error: {exc}"]}
                )
                continue
            successes += 1

        # ── Finalise record ───────────────────────────────────────────────────
        with transaction.atomic():
            record.status = BulkImportRecord.STATUS_COMPLETED
            record.success_count = successes
            record.failure_count = len(row_errors)
            record.row_errors = row_errors
            record.completed_at = tz.now()
            record.save()

            _audit(
                request,
                action="BULK_IMPORT_COMPLETED",
                entity_id=record.id,
                new_state={
                    "import_id": str(record.id),
                    "filename": uploaded.name,
                    "file_hash": file_hash,
                    "total": len(rows),
                    "success": successes,
                    "failures": len(row_errors),
                },
            )

        return _envelope(
            {
                "import_id": str(record.id),
                "total_rows": len(rows),
                "success_count": successes,
                "failure_count": len(row_errors),
                "row_errors": row_errors,
            },
            request_id=getattr(request, "request_id", ""),
        )


# ── Bulk Role Assignment  ─────────────────────────────────────────────


class BulkRoleAssignView(APIView):
    """``POST /api/v1/admin/users/bulk-roles`` — assign or revoke a role for
    multiple existing user profiles in one operation.

    Blueprint §1.2.4: "Bulk role assignment supported as a separate operation
    (existing profiles only, not creating new IAM accounts)."

    * Mutual-exclusion rules are enforced for each target user.
    * High-privilege roles follow the normal two-admin approval path —
      those users are skipped with an appropriate row-level error.
    * Each successful change emits a ``RoleChangeEvent`` and publishes
      a ``UserRoleChanged`` outbox event.
    """

    permission_classes = [IsAuthenticated, has_permission_with_step_up("users:manage")]

    @extend_schema(
        tags=["User Administration"],
        summary="Bulk assign or revoke a role across multiple profiles",
        operation_id="admin_users_bulk_roles",
        request=BulkRoleAssignSerializer,
        responses={
            200: _success_envelope(
                "BulkRoleResponse",
                {
                    "success_count": serializers.IntegerField(),
                    "failure_count": serializers.IntegerField(),
                    "errors": serializers.ListField(child=serializers.JSONField()),
                },
            ),
            400: _error_envelope("BulkRoleValidationError"),
            401: _error_envelope("BulkRoleUnauthorizedError"),
            403: _error_envelope("BulkRoleForbiddenError"),
        },
    )
    @transaction.atomic
    def post(self, request):
        from django.utils import timezone as tz
        from apps.users.models import (
            HIGH_PRIVILEGE_ROLES,
            Role,
            RoleChangeEvent,
            RoleMutualExclusion,
            UserProfile,
            UserRole,
        )
        from shared import keycloak_admin
        from shared.events import publish

        serializer = BulkRoleAssignSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action = serializer.validated_data["action"]
        role_name = serializer.validated_data["role"]
        user_ids = serializer.validated_data["user_ids"]
        effective_from = (
            serializer.validated_data.get("effective_from") or tz.now().date()
        )
        effective_to = serializer.validated_data.get("effective_to")
        reason = serializer.validated_data.get("reason", "Bulk operation")

        role_obj = Role.objects.get(name=role_name)

        # High-privilege roles must go through individual two-admin approval
        if role_name in HIGH_PRIVILEGE_ROLES and action == "assign":
            return _error_response(
                "HIGH_PRIVILEGE_ROLE",
                f"Role '{role_name}' requires individual two-administrator approval. "
                "Use POST /admin/users/{id}/roles/ for each user.",
                status.HTTP_400_BAD_REQUEST,
                getattr(request, "request_id", ""),
            )

        successes = 0
        errors = []

        for uid in user_ids:
            try:
                user = UserProfile.objects.get(pk=uid)
            except UserProfile.DoesNotExist:
                errors.append({"user_id": str(uid), "error": "Profile not found."})
                continue

            if action == "assign":
                conflict = RoleMutualExclusion.check_conflict(user, role_obj)
                if conflict:
                    errors.append(
                        {
                            "user_id": str(uid),
                            "email": user.email,
                            "error": (
                                f"Mutual exclusion: cannot assign '{role_name}' — "
                                "conflicts with '"
                                f"{conflict.role_a.name if conflict.role_b == role_obj else conflict.role_b.name}'."
                            ),
                        }
                    )
                    continue

                if UserRole.objects.filter(
                    user=user, role=role_obj, revoked_at__isnull=True
                ).exists():
                    errors.append(
                        {
                            "user_id": str(uid),
                            "email": user.email,
                            "error": f"Role '{role_name}' already assigned.",
                        }
                    )
                    continue

                if user.keycloak_sub:
                    try:
                        keycloak_admin.assign_client_role(
                            str(user.keycloak_sub), role_name
                        )
                    except Exception as exc:
                        logger.warning(
                            "bulk_role_assign: keycloak assign_client_role failed "
                            "user=%s role=%s: %s", uid, role_name, exc,
                        )
                        errors.append({
                            "user_id": str(uid),
                            "email": user.email,
                            "error": "IAM role mapping failed.",
                        })
                        continue

                UserRole.objects.create(
                    user=user,
                    role=role_obj,
                    effective_from=effective_from,
                    effective_to=effective_to,
                    assigned_by=request.user,
                )
                RoleChangeEvent.objects.create(
                    user=user,
                    role=role_obj,
                    change_type="assign",
                    actor=request.user,
                    reason=reason,
                )
                publish(
                    "UserRoleChanged",
                    {
                        "user_id": str(user.id),
                        "change_type": "assign",
                        "role": role_name,
                        "effective_from": str(effective_from),
                        "effective_to": str(effective_to) if effective_to else None,
                        "source": "bulk_operation",
                    },
                )

            else:  # revoke
                ur = UserRole.objects.filter(
                    user=user, role=role_obj, revoked_at__isnull=True
                ).first()
                if not ur:
                    errors.append(
                        {
                            "user_id": str(uid),
                            "email": user.email,
                            "error": f"Role '{role_name}' is not currently held.",
                        }
                    )
                    continue

                if user.keycloak_sub:
                    try:
                        keycloak_admin.remove_client_role(
                            str(user.keycloak_sub), role_name
                        )
                    except Exception as exc:
                        logger.warning(
                            "bulk_role_revoke: keycloak remove_client_role failed "
                            "user=%s role=%s: %s", uid, role_name, exc,
                        )
                        errors.append({
                            "user_id": str(uid),
                            "email": user.email,
                            "error": "IAM role removal failed.",
                        })
                        continue

                ur.revoked_at = tz.now()
                ur.revoke_reason = reason
                ur.save(update_fields=["revoked_at", "revoke_reason"])
                RoleChangeEvent.objects.create(
                    user=user,
                    role=role_obj,
                    change_type="revoke",
                    actor=request.user,
                    reason=reason,
                )
                publish(
                    "UserRoleChanged",
                    {
                        "user_id": str(user.id),
                        "change_type": "revoke",
                        "role": role_name,
                        "source": "bulk_operation",
                    },
                )

            successes += 1

        _audit(
            request,
            action="BULK_ROLE_OPERATION",
            new_state={
                "action": action,
                "role": role_name,
                "total": len(user_ids),
                "success": successes,
                "failures": len(errors),
                "effective_to": str(effective_to) if (effective_to and action == "assign") else None,
            },
        )

        return _envelope(
            {
                "success_count": successes,
                "failure_count": len(errors),
                "errors": errors,
            },
            request_id=getattr(request, "request_id", ""),
        )
